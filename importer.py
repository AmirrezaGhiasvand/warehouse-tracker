"""
importer.py
Reads the daily Excel file and syncs it into the SQLite database.

Expected columns (header row), matching the department's export format:
    نام_کالا          -> product name
    شماره_فني         -> part/technical code (unique key)
    مدل_به_کاررفته     -> model(s) used
    آدرس_در_انبار      -> warehouse location
"""

import os
import openpyxl
import db

# Map expected header names -> internal field names.
# If the department's export ever uses slightly different header text,
# add the alternate spelling here.
HEADER_MAP = {
    "نام_کالا": "name",
    "شماره_فني": "code",
    "مدل_به_کاررفته": "model",
    "آدرس_در_انبار": "location",
}


class ImportError_(Exception):
    pass


def _normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def import_excel_file(filepath, progress_callback=None):
    """
    Import an Excel file into the database.

    progress_callback(current, total) is called periodically if provided.

    Returns a dict summary:
        {filename, row_count, new_parts, location_changes, skipped}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ImportError_("The file appears to be empty.")

    header = [(_normalize(h)) for h in header]

    # Find column index for each expected field
    col_index = {}
    for idx, h in enumerate(header):
        if h in HEADER_MAP:
            col_index[HEADER_MAP[h]] = idx

    missing = [k for k in ("name", "code", "model", "location") if k not in col_index]
    if missing:
        raise ImportError_(
            "This file is missing expected column(s): " + ", ".join(missing) +
            "\n\nExpected headers: نام_کالا, شماره_فني, مدل_به_کاررفته, آدرس_در_انبار"
        )

    # Count total rows for progress (read_only sheets know dimensions)
    total_rows = ws.max_row - 1 if ws.max_row else 0

    filename = os.path.basename(filepath)

    conn = db.get_connection()
    cur = conn.cursor()

    row_count = 0
    new_parts = 0
    location_changes = 0
    skipped = 0

    for i, row in enumerate(rows_iter, start=1):
        code = _normalize(row[col_index["code"]]) if col_index["code"] < len(row) else ""
        name = _normalize(row[col_index["name"]]) if col_index["name"] < len(row) else ""
        model = _normalize(row[col_index["model"]]) if col_index["model"] < len(row) else ""
        location = _normalize(row[col_index["location"]]) if col_index["location"] < len(row) else ""

        if not code:
            skipped += 1
            continue

        is_new, changed = db.upsert_part(cur, code, name, model, location, filename)
        row_count += 1
        if is_new:
            new_parts += 1
        elif changed:
            location_changes += 1

        if progress_callback and i % 200 == 0:
            progress_callback(i, total_rows)

    db.log_upload(cur, filename, row_count, new_parts, location_changes)
    conn.commit()
    conn.close()

    if progress_callback:
        progress_callback(total_rows, total_rows)

    return {
        "filename": filename,
        "row_count": row_count,
        "new_parts": new_parts,
        "location_changes": location_changes,
        "skipped": skipped,
    }
